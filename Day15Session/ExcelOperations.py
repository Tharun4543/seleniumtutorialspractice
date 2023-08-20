import openpyxl as op
workbook = op.load_workbook("D:\\personalfiles\\Seat_Information_2022.xlsx")
sheet = workbook["2LAKHS_SEAT_INFO"]
max_rows = sheet.max_row
max_columns = sheet.max_column

for row in range(1,max_rows):
    for col in range(1,max_columns):
        print(sheet.cell(row,col).value, end = " ")
    print(" ")


